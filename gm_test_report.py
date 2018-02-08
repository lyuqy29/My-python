# -*- coding: utf-8 -*-

import openpyxl
import time
import datetime
import os
from openpyxl import Workbook
from openpyxl.chart import BarChart3D,Series,Reference,PieChart,LineChart,PieChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.layout import Layout, ManualLayout


##Date define
now_time = datetime.datetime.now()
today_time = now_time + datetime.timedelta()
yes_time = now_time + datetime.timedelta(days=-1)


date= today_time.strftime('%Y%m%d')
date2= today_time.strftime('%Y/%m/%d')
date3= yes_time.strftime('%Y%m%d')

#Version define
server_version="rdb-1.1.5.6"
vehicle_version="rdb-1.1.5.5"

##Column range

column_start=1
column_end=18


# xfile = openpyxl.load_workbook("process_status_GM_"+date3+".xlsx")
if os.path.exists('/Users/test/PycharmProjects/gm_test/'+"process_status_GM_"+date3+".xlsx"):
    xfile = openpyxl.load_workbook(
        filename='/Users/test/PycharmProjects/gm_test/' + "process_status_GM_" + date3 + ".xlsx")
else:
    xfile = openpyxl.load_workbook(filename='/Users/test/PycharmProjects/gm_test/'+"process_status_GM"+".xlsx")

sheet = xfile.get_sheet_by_name("daily statistic 5th Round")

row=sheet.max_row
new_data_row=row

print new_data_row,row

#Input new data for new colomn

# temp=[1,2,3,4,5,6,7,8,"Rdb-1.1.1.35","Rdb-1.1.1.35","Test"]
f= os.popen("sh get_data.sh") #Call script to get the data
f.close
all=f.readlines()#obtain the temp data

data=[]


for temp_data in all:
    b=int(temp_data.strip()) #Transform into int type
    data.append(b)

print data



server_version_list = []
vehicle_version_list = []

# Get the list of server version
for tem_row in range(2, row+1):
    if sheet.cell(row=tem_row, column=16).value:
        server_version_list.append(sheet.cell(row=tem_row, column=16).value)
    pass

# Get the list of vehicle version
for tem_row in range(2, row+1):
    if sheet.cell(row=tem_row, column=17).value:
        vehicle_version_list.append(sheet.cell(row=tem_row, column=17).value)
    pass


flag = 0
# input new version if need
if server_version not in server_version_list or vehicle_version not in vehicle_version_list:
    sheet.cell(row=new_data_row, column=16).value = server_version
    sheet.cell(row=new_data_row, column=17).value = vehicle_version
    server_version_list.append(server_version)
    vehicle_version_list.append(vehicle_version)
    flag =1


calculate_row = 0
for tem_row in range(2, row+1):
    if sheet.cell(row=tem_row, column=16).value == server_version_list[-1] and sheet.cell(row=tem_row,
                                                                                          column=17).value == \
            vehicle_version_list[-1]:
        calculate_row = tem_row
        break
    pass
print "calculate_row",calculate_row

#Fill the original data
sheet.cell(row=new_data_row,column=1).value=date2
t=0
skip_col=2
for col in range(column_start+1,column_end):
    if skip_col in [7,8,10,11,14,16,17]:
        skip_col=skip_col+1
        continue
    if skip_col == 6:
        sheet.cell(row=new_data_row, column=col).value = sum(data[0:4])
        skip_col = skip_col + 1
        continue
    sheet.cell(row=new_data_row,column=col).value=data[t]
    # print col,data[t],skip_col
    t=t+1
    skip_col=skip_col+1

#Fill the processed_total data
processed_total = 0
for row2 in range(calculate_row,row+1):
    print row2,processed_total
    processed_total = processed_total+ sheet.cell(row=row2,column=6).value
   # print row2,processed_total
sheet.cell(row=new_data_row,column=7).value=processed_total

#
#Fill the uploads daily data
print "flag",flag
if flag == 1:
    uploads_daily = data[4]
    print uploads_daily
else:
    uploads_daily = data[4] - sheet.cell(row=new_data_row - 1, column=9).value
    print uploads_daily
sheet.cell(row=new_data_row, column=8).value = uploads_daily

#Fill the Remaining
sheet.cell(row=new_data_row, column=10).value = data[7]-processed_total

#Fill the failed
sheet.cell(row=new_data_row, column=11).value = processed_total - data[4]

#Fill the T2 failed
sheet.cell(row=new_data_row, column=14).value= data[4] - sum(data[5:7])

empty_list=[]

#Fill total part:
empty_list.append("Total")

# #Fill Car total
Car1_total = 0
Car2_total = 0
Car3_total = 0
Car4_total = 0
for car_row in range(calculate_row,new_data_row+1):
    Car1_total = Car1_total+ sheet.cell(row=car_row,column=2).value
    Car2_total = Car2_total + sheet.cell(row=car_row, column=3).value
    Car3_total = Car3_total + sheet.cell(row=car_row, column=4).value
    Car4_total = Car4_total + sheet.cell(row=car_row, column=5).value

empty_list.append(Car1_total)
empty_list.append(Car2_total)
empty_list.append(Car3_total)
empty_list.append(Car4_total)

#Fill the total process
empty_list.append(processed_total)

#Fill the total process
empty_list.append(processed_total)

#Fill the daily total upload
empty_list.append(data[4])

#Fill the total upload
empty_list.append(data[4])

#Fill the total remain
empty_list.append(data[7]-processed_total)

#Fill the total failed
empty_list.append(processed_total-data[4])

#Fill the total T2 success
empty_list.append(data[5])

#Fill the total T2 waiting
empty_list.append(data[6])

#Fill the T2 failed
empty_list.append(data[4] - sum(data[5:7]))

#Fill the total RTV
empty_list.append(data[7])
sheet.append(empty_list)


pie= PieChart3D()
labels = Reference(sheet,min_row=1,min_col=9,max_col=11)
data = Reference(sheet,min_row=10,min_col=9,max_col=11)

pie.add_data(data,data,titles_from_data=False)
pie.set_categories(labels=labels)
pie.title="Report 2"
slice=DataPoint(idx=0,explosion=10)
pie.series[0].data_points = [slice]

sheet.add_chart(pie,'k'+str(new_data_row+4))


c1 = LineChart()
c1.title = "Report 1"
c1.style= 10
c1.y_axis.title = "RTV"
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.x_axis.title = "Date"

data1= Reference(sheet,min_col=7,min_row=1,max_row=new_data_row,max_col=15)
date1 = Reference(sheet,min_col=1,min_row=2,max_row=new_data_row)
c1.add_data(data1,titles_from_data=True)
c1.set_categories(date1)

c1.width=22
c1.height=12

xfile.save(filename='/Users/test/PycharmProjects/gm_test/'+"process_status_GM_"+date+".xlsx")
xfile.close()




